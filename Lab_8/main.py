import argparse
import csv
import sys
import openpyxl
from openpyxl.styles import Font


def run():
    argument = argparse.ArgumentParser()
    argument.add_argument('file_name', type=str, help="File need a name")
    argument.add_argument('-o', help="For saving result use this command")
    my_args = argument.parse_args()
    data = my_args.file_name
    if data[-4:] != '.csv':
        print('file name is wrong')
        sys.exit(0)
    else:
        data_name = read_log(data)
        best_seller(data_name)
        top_played(data_name)
        sum_of_total_games(data_name)
        average_time(data_name)

    if 'o' in my_args and my_args.o is not None:
        my_file = my_args.o
        excel_file(data_name, my_file)
    else:
        print(best_seller(data_name))
        print(top_played(data_name))
        print(sum_of_total_games(data_name))
        print(average_time(data_name))


class read_content:
    def __init__(self, lines):
        self.id = lines['user_id']
        self.name = lines['name_of_the_game']
        self.behavior = lines['behavior_name']
        self.hours = float(lines['hours_if_behavior_is_play'])

    def __str__(self):
        return (f'Id number: {self.id} '
                f'Name of the game: {self.name} '
                f'Is it purchase or play: {self.behavior} '
                f'How many hours person spent time: {self.hours}')

    def game_name(self):
        return self.name

    def time(self):
        return self.hours

    def behaviour_name(self):
        return self.behavior


def read_log(data_name):
    data_list = []
    try:
        with open(data_name) as csv_file:  # exception is not working well check it
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                data_list.append(read_content(row))

        return data_list
    except FileNotFoundError:
        print('File is empty')
        sys.exit(0)


def best_seller(file):
    count = 0
    name_list = {}
    name_list1 = {}
    for lines in file:
        if 'play' not in lines.behaviour_name():
            if lines.game_name() not in name_list:
                name_list[lines.game_name()] = 1
            else:
                name_list[lines.game_name()] += 1
    sorted_values = sorted(name_list.values())
    name_list2 = {}
    for i in sorted_values:
        for k in name_list.keys():
            if name_list[k] == i:
                name_list2[k] = name_list[k]
                break
    for elem in name_list2:
        count += 1
        if count >= (len(name_list2) - 9):
            name_list1[elem] = name_list2[elem]
            # print(f'Best selling games:{elem} , times: {name_list2[elem]}')
    return name_list1


def top_played(file):
    count = 0
    name_list = {}
    game_name = ''
    for lines in file:
        if 'purchase' not in lines.behaviour_name():
            if lines.game_name() not in name_list:
                name_list[lines.game_name()] = lines.time()
            else:
                name_list[lines.game_name()] += lines.time()
    for elem in name_list:
        if count < float(name_list[elem]):
            count = name_list[elem]
            game_name = elem
    # print(f'\nTop played game: {game_name}, hours: {round(count)}')
    return [game_name, count]


def sum_of_total_games(file):
    count = 0
    list_of_games = {}
    for lines in file:
        if lines.game_name() not in list_of_games:
            list_of_games[lines.game_name()] = 1
            count += 1
    # print(f'\nTotal games on the market:{count}')
    return count


def average_time(file):
    count = 0
    sum_of_hours = 0
    for lines in file:
        if 'purchase' not in lines.behaviour_name():
            sum_of_hours = sum_of_hours + float(lines.time())
            count += 1
    average = sum_of_hours / count
    # print(f'\nAverage hours playing on steam: {int(average)}')
    return average


def excel_file(file, name_file):
    font_type = Font(name='Calibri', size=11, bold=True, italic=True,
                     vertAlign=None, underline='single', color='ff0000')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Steam Data Set"
    average_hour = sheet.cell(row=1, column=1)
    average_hour.value = "Avg hours played"
    average_hour.font = font_type
    average_hour_inside = sheet.cell(row=2, column=1)
    average_hour_inside.value = average_time(file)

    sum_of_games = sheet.cell(row=1, column=3)
    sum_of_games.value = "Total num. of games"
    sum_of_games.font = font_type
    sum_of_games_inside = sheet.cell(row=2, column=3)
    sum_of_games_inside.value = sum_of_total_games(file)

    bestseller = sheet.cell(row=1, column=6)
    bestseller.value = "Bestsellers:"
    bestseller.font = font_type
    new_file = best_seller(file)
    i = 2
    for obj in new_file:
        obj_cell = sheet.cell(row=i, column=6)
        obj_cell.value = obj
        i += 1

    top_played_game = sheet.cell(row=1, column=9)
    top_played_game.value = "Most played game"
    top_played_game.font = font_type
    top_played_game_data = sheet.cell(row=2, column=9)
    top_played_game_data.value = top_played(file)[0]
    hours = sheet.cell(row=1, column=10)
    hours.value = "Hours"
    hours.font = font_type
    hours_data = sheet.cell(row=2, column=10)
    hours_data.value = top_played(file)[1]

    workbook.save(name_file) #its not working well


if __name__ == '__main__':
    run()
